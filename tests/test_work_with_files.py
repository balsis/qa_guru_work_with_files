import csv
from io import TextIOWrapper
from pypdf import PdfReader
from openpyxl import load_workbook


class TestFiles:
    def test_csv_file(self, csv_file):
        csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig')))
        second_row = csvreader[1]
        assert second_row[0] == 'Jack'

    def test_pdf_file(self, pdf_file):
        reader = PdfReader(pdf_file)
        number_of_pages = len(reader.pages)
        assert number_of_pages == 1
        page = reader.pages[0]
        text = page.extract_text()
        assert "Lorem ipsum dolor sit amet" in text

    def test_xlsx_file(self, xlsx_file):
        work_book = load_workbook(xlsx_file)
        sheet = work_book.active
        cell_value = sheet.cell(row = 5, column = 5).value
        max_row = sheet.max_row
        max_column = sheet.max_column
        assert cell_value == "United States"
        assert max_column == 8
        assert max_row == 10
